from io import BytesIO

from openpyxl import load_workbook

from domain.errors import ServiceError
from schemas.admin import CourseImportRow
from services.course_management_service import CourseManagementService


REQUIRED_COLUMNS = (
    "course_code", "course_name", "department", "curriculum_name", "curriculum_year",
    "recommended_year", "recommended_semester", "requirement_type",
)
OPTIONAL_COLUMNS = ("prerequisites", "syllabus")
MAX_UPLOAD_BYTES = 5 * 1024 * 1024


class CourseImportService:
    def __init__(self, course_management_service=None):
        self.management = course_management_service or CourseManagementService()

    async def preview(self, upload):
        if not (upload.filename or "").lower().endswith(".xlsx"):
            raise ServiceError(415, "Upload an .xlsx Excel file.")
        content = await upload.read(MAX_UPLOAD_BYTES + 1)
        if len(content) > MAX_UPLOAD_BYTES:
            raise ServiceError(413, "Excel file exceeds the 5MB limit.")
        try:
            worksheet = load_workbook(BytesIO(content), read_only=True, data_only=True).active
        except Exception as exc:
            raise ServiceError(422, "The uploaded file is not a valid .xlsx workbook.") from exc
        header_row_number, positions = None, None
        for candidate_number, candidate in enumerate(worksheet.iter_rows(max_row=10, values_only=True), start=1):
            headers = [self._text(value).lower() for value in candidate]
            candidate_positions = {name: index for index, name in enumerate(headers) if name}
            if all(name in candidate_positions for name in REQUIRED_COLUMNS):
                header_row_number, positions = candidate_number, candidate_positions
                break
        missing = [name for name in REQUIRED_COLUMNS if positions is None or name not in positions]
        if missing:
            raise ServiceError(422, f"Missing required columns: {', '.join(missing)}")

        rows, seen_codes = [], set()
        for row_number, row in enumerate(worksheet.iter_rows(min_row=header_row_number + 1, values_only=True), start=header_row_number + 1):
            values = {name: self._text(row[index]) if index < len(row) else "" for name, index in positions.items()}
            if not any(values.values()):
                break
            errors = []
            code = values["course_code"].upper()
            if code in seen_codes:
                errors.append("duplicate course_code in this file")
            seen_codes.add(code)
            normalized = {
                "row_number": row_number, "course_code": code, "course_name": values["course_name"],
                "department": values["department"], "curriculum_name": values["curriculum_name"],
                "curriculum_year": self._integer(values["curriculum_year"]),
                "recommended_year": self._integer(values["recommended_year"]),
                "recommended_semester": values["recommended_semester"],
                "requirement_type": values["requirement_type"].upper() or "REQUIRED",
                "prerequisites": values.get("prerequisites") or None, "syllabus": values.get("syllabus") or None,
            }
            try:
                CourseImportRow(**normalized)
            except Exception as exc:
                errors.extend(item["msg"] for item in exc.errors())
            rows.append({**normalized, "errors": errors})
        if not rows:
            raise ServiceError(422, "The worksheet has no course rows.")
        conn = self.management.connection_factory()
        try:
            for row in rows:
                if row["errors"]:
                    row["operation"] = "invalid"
                    continue
                row["operation"] = self._operation_for(conn, row)
        finally:
            conn.close()
        return {"rows": rows, "valid_count": sum(not item["errors"] for item in rows), "invalid_count": sum(bool(item["errors"]) for item in rows)}

    def confirm(self, rows, admin, ip_address=None):
        conn = self.management.connection_factory()
        try:
            outcomes = []
            for row in rows:
                existing = self.management.courses.find_course_by_code_for_update(conn, row.course_code.strip().upper())
                if existing and self._is_exact_match(self.management.courses.get_import_state(conn, row.course_code.strip().upper()), row):
                    outcomes.append({"row_number": row.row_number, "course_id": existing["course_id"], "operation": "skipped"})
                    continue
                curriculum_id = self.management.courses.find_or_create_curriculum(
                    conn, row.curriculum_name.strip(), row.curriculum_year, row.department.strip()
                )
                data = type("ImportedCourse", (), {
                    "course_code": row.course_code, "course_name": row.course_name,
                    "department": row.department, "prerequisites": row.prerequisites,
                    "syllabus": row.syllabus, "teaching_format": None, "workload": None,
                    "assessment": None,
                })()
                if existing:
                    course_id = existing["course_id"]
                    self.management.courses.update_imported_course(conn, course_id, data)
                    operation = "updated"
                else:
                    course_id = self.management.courses.create_course(conn, data)
                    operation = "created"
                mapping = type("Mapping", (), {
                    "curriculum_id": curriculum_id, "recommended_year": row.recommended_year,
                    "recommended_semester": row.recommended_semester, "requirement_type": row.requirement_type,
                })()
                self.management.courses.replace_mappings(conn, course_id, [mapping])
                self.management.audit.create(conn, admin["user_id"], "MANAGE_COURSE", course_id, ip_address)
                outcomes.append({"row_number": row.row_number, "course_id": course_id, "operation": operation})
            conn.commit()
            return {"imported_count": len(outcomes), "results": outcomes}
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

    @staticmethod
    def _text(value):
        return "" if value is None else str(value).strip()

    @staticmethod
    def _integer(value):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return 0

    def _operation_for(self, conn, row):
        state = self.management.courses.get_import_state(conn, row["course_code"])
        if state is None:
            return "create"
        return "skip" if self._is_exact_match(state, row) else "update"

    @staticmethod
    def _is_exact_match(state, row):
        if state is None:
            return False
        value = (lambda name: row[name]) if isinstance(row, dict) else (lambda name: getattr(row, name))
        course, mappings = state
        same_course = (
            course["course_name"].strip() == value("course_name").strip()
            and course["department"].strip() == value("department").strip()
            and (course["prerequisites"] or "").strip() == (value("prerequisites") or "").strip()
            and (course["syllabus"] or "").strip() == (value("syllabus") or "").strip()
        )
        if not same_course or len(mappings) != 1:
            return False
        mapping = mappings[0]
        return (
            mapping["curriculum_name"].strip() == value("curriculum_name").strip()
            and mapping["academic_year"] == value("curriculum_year")
            and mapping["recommended_year"] == value("recommended_year")
            and mapping["recommended_semester"].strip() == value("recommended_semester").strip()
            and mapping["requirement_type"] == value("requirement_type")
        )
