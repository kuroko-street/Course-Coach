import csv
from io import BytesIO, StringIO

import psycopg2
from openpyxl import load_workbook

from db import get_connection
from domain.errors import ServiceError
from repositories.audit_log_repository import AuditLogRepository
from repositories.student_import_repository import StudentImportRepository
from schemas.admin import StudentEnrollmentImportRow


REQUIRED_COLUMNS = (
    "student_number", "email", "course_code", "academic_year", "semester", "section",
)
MAX_UPLOAD_BYTES = 5 * 1024 * 1024


class StudentImportService:
    def __init__(self, connection_factory=get_connection):
        self.connection_factory = connection_factory
        self.students = StudentImportRepository()
        self.audit = AuditLogRepository()

    def list_students(self):
        conn = self.connection_factory()
        try:
            return {"students": self.students.list_students(conn)}
        finally:
            conn.close()

    async def preview(self, upload):
        filename = (upload.filename or "").lower()
        if not filename.endswith((".csv", ".xlsx")):
            raise ServiceError(415, "Upload a .csv or .xlsx file.")
        content = await upload.read(MAX_UPLOAD_BYTES + 1)
        if len(content) > MAX_UPLOAD_BYTES:
            raise ServiceError(413, "Student import file exceeds the 5MB limit.")

        raw_rows = self._read_csv(content) if filename.endswith(".csv") else self._read_xlsx(content)
        if len(raw_rows) > 2000:
            raise ServiceError(413, "Student import supports at most 2,000 rows per file.")
        rows, seen = [], set()
        number_emails, email_numbers = {}, {}
        for row_number, values in raw_rows:
            normalized = {
                "row_number": row_number,
                "student_number": self._student_number(values.get("student_number")),
                "email": self._text(values.get("email")).casefold(),
                "course_code": self._text(values.get("course_code")).upper(),
                "academic_year": self._integer(values.get("academic_year")),
                "semester": self._text(values.get("semester")),
                "section": self._section(values.get("section")),
            }
            errors = []
            identity = (
                normalized["student_number"], normalized["course_code"],
                normalized["academic_year"], normalized["semester"], normalized["section"],
            )
            if identity in seen:
                errors.append("duplicate enrollment in this file")
            seen.add(identity)
            known_email = number_emails.setdefault(normalized["student_number"], normalized["email"])
            if known_email != normalized["email"]:
                errors.append("student_number uses more than one email in this file")
            known_number = email_numbers.setdefault(normalized["email"], normalized["student_number"])
            if known_number != normalized["student_number"]:
                errors.append("email uses more than one student_number in this file")
            try:
                StudentEnrollmentImportRow(**normalized)
            except Exception as exc:
                errors.extend(item["msg"] for item in exc.errors())
            rows.append({**normalized, "errors": errors})

        if not rows:
            raise ServiceError(422, "The file has no student enrollment rows.")

        conn = self.connection_factory()
        try:
            for row in rows:
                if row["errors"]:
                    row["operation"] = "invalid"
                    continue
                state = self._state(conn, row)
                row["errors"].extend(state["errors"])
                row["operation"] = "invalid" if row["errors"] else state["operation"]
                row["course_name"] = state.get("course_name")
                row["google_linked"] = state.get("google_linked", False)
        finally:
            conn.close()

        return {
            "rows": rows,
            "valid_count": sum(not row["errors"] for row in rows),
            "invalid_count": sum(bool(row["errors"]) for row in rows),
        }

    def confirm(self, rows, admin, ip_address=None):
        conn = self.connection_factory()
        try:
            results = []
            for row in rows:
                data = row.model_dump()
                state = self._state(conn, data, lock=True)
                if state["errors"]:
                    raise ServiceError(409, f"Row {row.row_number}: {state['errors'][0]}")

                user = state["user"]
                if user is None:
                    user = self.students.create_student(conn, row.student_number, row.email)
                elif not user["student_number"]:
                    user = self.students.attach_student_number(conn, user["user_id"], row.student_number)

                enrollment = self.students.create_enrollment(
                    conn, user["user_id"], state["course"]["course_id"],
                    row.academic_year, row.semester, row.section,
                )
                if enrollment is None:
                    existing = state["enrollment"] or self.students.enrollment_exists(
                        conn, user["user_id"], state["course"]["course_id"],
                        row.academic_year, row.semester, row.section,
                    )
                    operation, enrollment_id = "skipped", existing["enrollment_id"]
                else:
                    operation, enrollment_id = state["operation"], enrollment["enrollment_id"]
                    self.audit.create(
                        conn, admin["user_id"], "IMPORT_ENROLLMENT", enrollment_id, ip_address
                    )
                results.append({
                    "row_number": row.row_number,
                    "user_id": user["user_id"],
                    "enrollment_id": enrollment_id,
                    "operation": operation,
                })
            conn.commit()
            return {
                "processed_count": len(results),
                "created_count": sum(item["operation"] != "skipped" for item in results),
                "skipped_count": sum(item["operation"] == "skipped" for item in results),
                "results": results,
            }
        except ServiceError:
            conn.rollback()
            raise
        except psycopg2.IntegrityError as exc:
            conn.rollback()
            raise ServiceError(409, "Student number or email conflicts with existing data.") from exc
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

    def _state(self, conn, row, lock=False):
        value = (lambda name: row[name]) if isinstance(row, dict) else (lambda name: getattr(row, name))
        number, email = value("student_number"), value("email")
        by_number, by_email = self.students.identity_matches(conn, number, email, lock=lock)
        errors = []
        if by_number and by_email and by_number["user_id"] != by_email["user_id"]:
            errors.append("student_number and email belong to different accounts")
        elif by_number and by_number["email"].casefold() != email.casefold():
            errors.append("student_number is already assigned to another email")
        elif by_email and by_email["student_number"] not in (None, number):
            errors.append("email is already assigned to another student_number")

        user = by_number or by_email
        if user and user["role"] != "STUDENT":
            errors.append("email belongs to an administrator account")

        course = self.students.find_course(conn, value("course_code"))
        if course is None:
            errors.append("course_code does not exist")

        enrollment = None
        if user and course and not errors:
            enrollment = self.students.enrollment_exists(
                conn, user["user_id"], course["course_id"], value("academic_year"),
                value("semester"), value("section"),
            )

        if errors:
            operation = "invalid"
        elif enrollment:
            operation = "skip"
        elif user is None:
            operation = "create_student"
        elif not user["student_number"]:
            operation = "link_existing_user"
        else:
            operation = "add_enrollment"
        return {
            "errors": errors,
            "operation": operation,
            "user": user,
            "course": course,
            "course_name": course["course_name"] if course else None,
            "enrollment": enrollment,
            "google_linked": bool(user and user["google_linked"]),
        }

    def _read_csv(self, content):
        try:
            reader = csv.DictReader(StringIO(content.decode("utf-8-sig")))
        except UnicodeDecodeError as exc:
            raise ServiceError(422, "CSV file must use UTF-8 encoding.") from exc
        headers = [self._text(name).casefold() for name in (reader.fieldnames or [])]
        missing = [name for name in REQUIRED_COLUMNS if name not in headers]
        if missing:
            raise ServiceError(422, f"Missing required columns: {', '.join(missing)}")
        return [
            (row_number, {self._text(key).casefold(): value for key, value in row.items() if key})
            for row_number, row in enumerate(reader, start=2)
            if any(self._text(value) for value in row.values())
        ]

    def _read_xlsx(self, content):
        try:
            sheet = load_workbook(BytesIO(content), read_only=True, data_only=True).active
        except Exception as exc:
            raise ServiceError(422, "The uploaded file is not a valid .xlsx workbook.") from exc
        header_number, positions = None, None
        for number, candidate in enumerate(sheet.iter_rows(max_row=10, values_only=True), start=1):
            headers = [self._text(value).casefold() for value in candidate]
            found = {name: index for index, name in enumerate(headers) if name}
            if all(name in found for name in REQUIRED_COLUMNS):
                header_number, positions = number, found
                break
        missing = [name for name in REQUIRED_COLUMNS if positions is None or name not in positions]
        if missing:
            raise ServiceError(422, f"Missing required columns: {', '.join(missing)}")
        rows = []
        for number, row in enumerate(
            sheet.iter_rows(min_row=header_number + 1, values_only=True), start=header_number + 1
        ):
            values = {name: row[index] if index < len(row) else None for name, index in positions.items()}
            if any(self._text(value) for value in values.values()):
                rows.append((number, values))
        return rows

    @staticmethod
    def _text(value):
        return "" if value is None else str(value).strip()

    @staticmethod
    def _integer(value):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return 0

    @classmethod
    def _student_number(cls, value):
        text = cls._text(value)
        if text.endswith(".0") and text[:-2].isdigit():
            return text[:-2]
        return text

    @classmethod
    def _section(cls, value):
        text = cls._text(value)
        if text.endswith(".0") and text[:-2].isdigit():
            text = text[:-2]
        return text.zfill(3) if text.isdigit() and len(text) < 3 else text
